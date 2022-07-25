#!/usr/bin/env python3

import argparse
from datetime import datetime, time
import os
from pathlib import Path
import sys

from bs4 import BeautifulSoup
from bs4.element import Tag
import bokeh
from bokeh.models import HoverTool, DataRange1d
from bokeh.plotting import figure
from bokeh.embed import components
import numpy as np
from openpyxl import load_workbook

def property_table(html_tree, names, values):
    table = html_tree.new_tag('table')

    assert len(names) == len(values)
    for i in range(len(names)):
        name  = names[i]
        value = values[i]

        tr = html_tree.new_tag('tr')
        td_name = html_tree.new_tag('td')
        td_name.string = name
        tr.append(td_name)
        td_value = html_tree.new_tag('td')
        td_value.string = value
        tr.append(td_value)

        table.append(tr)

    return table

def hyperlink(html_tree, text, href):
    a = html_tree.new_tag('a', href = href)
    a.string = text
    return a

def calc_total_g(x_g, y_g, z_g):
    return np.sqrt( np.array(x_g)**2 + np.array(y_g)**2 + np.array(z_g)**2 )

class MSR175WorkbookLoadError(Exception):
    '''Indicates that the workbook has an error.'''
    
    def __init__(self, xlsx_path, sheetname, cell_address, message):
        xlsx_filename = Path(xlsx_path).name
        msg = f'[{xlsx_filename}]\'{sheetname}\'!{cell_address}: {message}'
        super().__init__(msg)

        self.__xlsx_path    = xlsx_path
        self.__sheetname    = sheetname
        self.__cell_address = cell_address
        self.__message      = message

    @property
    def xlsx_path(self):
        return self.__xlsx_path

    @property
    def sheetname(self):
        return self.__sheetname
    
    @property
    def cell_address(self):
        return self.__cell_address

    @property
    def message(self):
        return self.__message

class MSR175WorksheetParseError(Exception):
    '''Indicates that the worksheet is not in an expected format.'''
    
    def __init__(self, cell_address, message):
        super().__init__(message)
        self.__cell_address = cell_address
        self.__message      = message

    @property
    def cell_address(self):
        return self.__cell_address

    @property
    def message(self):
        return self.__message

class MSR175ShockEvent:
    def __init__(self, event_id, timestamp, sampling_period_ms, x_g, y_g, z_g):
        self.__xlsx_path = None
        self.__event_id  = event_id
        self.__timestamp = timestamp
        self.__sampling_period_ms = sampling_period_ms
        self.__x_g = x_g
        self.__y_g = y_g
        self.__z_g = z_g

        assert len(x_g) == len(y_g)
        assert len(y_g) == len(z_g)

        # Calculate total acceleration.
        self.__total_g = calc_total_g(x_g, y_g, z_g)

        # Calculate power spectrum
        n = len(x_g)
        ps_x_g2 = np.abs(np.fft.fft(x_g))**2
        ps_y_g2 = np.abs(np.fft.fft(y_g))**2
        ps_z_g2 = np.abs(np.fft.fft(z_g))**2
        freq_Hz = np.fft.fftfreq(n, sampling_period_ms / 1000.0)

        self.__ps_x_g2 = ps_x_g2[0:int(n/2)]
        self.__ps_y_g2 = ps_y_g2[0:int(n/2)]
        self.__ps_z_g2 = ps_z_g2[0:int(n/2)]
        self.__freq_Hz = freq_Hz[0:int(n/2)]

    @property
    def event_id(self):
        return self.__event_id

    @property
    def timestamp(self):
        return self.__timestamp

    @property
    def sampling_period_ms(self):
        return self.__sampling_period_ms

    @property
    def sampling_frequency_Hz(self):
        return 1000.0 / self.__sampling_period_ms

    @property
    def duration_ms(self):
        return self.sampling_period_ms * self.n

    @property
    def n(self):
        return len(self.x_g)

    @property
    def x_g(self):
        return self.__x_g

    @property
    def y_g(self):
        return self.__y_g

    @property
    def z_g(self):
        return self.__z_g

    @property
    def total_g(self):
        return self.__total_g

    @property
    def max_g(self):
        return max(self.__total_g)

    @property
    def t_ms(self):
        return [self.sampling_period_ms * i for i in range(self.n)]

    @property
    def power_spectrum_x_g2(self):
        return self.__ps_x_g2

    @property
    def power_spectrum_y_g2(self):
        return self.__ps_y_g2
    
    @property
    def power_spectrum_z_g2(self):
        return self.__ps_z_g2

    @property
    def power_spectrum_freq_Hz(self):
        return self.__freq_Hz
    
    @property
    def xlsx_path(self):
        return Path(self.__xlsx_path)

    @property
    def xlsx_filename(self):
        return self.xlsx_path.name

    @property
    def html_id(self):
        # TODO: replace characters that are not allowed as a part of HTML id
        return f'{self.xlsx_filename}:{self.event_id}'

    def time_series_plot(self,
                         width     = 700,
                         height    = 350,
                         acc_min_g = None,
                         acc_max_g = None,
                         t_min_ms  = None,
                         t_max_ms  = None):
        
        plot = figure(plot_width  = width,
                      plot_height = height,
                      x_range = (0 if t_min_ms is None else t_min_ms,
                                 self.duration_ms if t_max_ms is None else t_max_ms),
                      x_axis_label = 'Time [ms]',
                      y_range = DataRange1d(start = acc_min_g, end = acc_max_g),
                      y_axis_label = 'Acceleration [g]',
                      toolbar_location = 'above')

        data = {
            't_ms'   : self.t_ms,
            'x_g'    : self.x_g,
            'y_g'    : self.y_g,
            'z_g'    : self.z_g,
            'total_g': self.total_g,
        }

        for label, y, color in (('X', 'x_g', 'red'),
                                ('Y', 'y_g', 'blue'),
                                ('Z', 'z_g', 'green'),
                                ('Total', 'total_g', 'orange')):
            plot.line(source = data, x = 't_ms', y = y, color = color, legend_label = label)
            plot.circle(source = data, size = 10, x = 't_ms', y = y, color = color,
                        alpha = 0.0, hover_color = color, hover_alpha = 1.0)

        plot.legend.location = 'bottom_right'
        tooltips = [('t', '@t_ms{0.00000} ms'),
                    ('x', '@x_g g'),
                    ('y', '@y_g g'),
                    ('z', '@z_g g'),
                    ('total', '@total_g g')]
        plot.add_tools(HoverTool(tooltips = tooltips))
        return plot

    def power_spectrum_plot(self,
                            width     = 700,
                            height    = 350,
                            ps_min_g2 = None,
                            ps_max_g2 = None):

        plot = figure(plot_width   = width,
                      plot_height  = height,
                      x_range      = (0, self.sampling_frequency_Hz / 2),
                      x_axis_label = 'Frequency [Hz]',
                      y_range = DataRange1d(start = ps_min_g2, end = ps_max_g2),
                      y_axis_label = 'Power Spectrum [g²]',
                      y_axis_type  = 'log',
                      toolbar_location = 'above')

        data = {
            'freq_Hz': self.power_spectrum_freq_Hz,
            'ps_x_g2': self.power_spectrum_x_g2,
            'ps_y_g2': self.power_spectrum_y_g2,
            'ps_z_g2': self.power_spectrum_z_g2,
        }

        for label, y, color in (('X', 'ps_x_g2', 'red'),
                                ('Y', 'ps_y_g2', 'blue'),
                                ('Z', 'ps_z_g2', 'green')):
            plot.line(source = data, x = 'freq_Hz', y = y, color = color, legend_label = label)
            plot.circle(source = data, size = 10, x = 'freq_Hz', y = y, color = color,
                        alpha = 0.0, hover_color = color, hover_alpha = 1.0)

        plot.legend.location = 'top_right'
        tooltips = [('Frequency', '@freq_Hz{0} Hz'),
                    ('X', '@ps_x_g2 g²'),
                    ('Y', '@ps_y_g2 g²'),
                    ('Z', '@ps_z_g2 g²')]
        plot.add_tools(HoverTool(tooltips = tooltips))
        return plot
    
    @staticmethod
    def validate_cell(worksheet, cell_address, expected_value):
        value = worksheet[cell_address].value
        if value != expected_value:
            message = f'Expected value was "{expected_value}", but it was "{value}".'
            raise MSR175WorksheetParseError(cell_address, message)

    @staticmethod
    def parse_date(worksheet, cell_address):
        value = worksheet[cell_address].value
        try:
            return datetime.strptime(value, '%y-%m-%d')
        except ValueError as e:
            message = f'Expected format was "%y-%m-%d" (e.g. 22-01-31), but it was "{value}". {str(e)}'
            raise MSR175WorksheetParseError(cell_address, message)
            
    @staticmethod
    def parse_time(worksheet, cell_address):
        value = worksheet[cell_address].value
        if isinstance(value, time):
            return value
        else:
            raise MSR175WorksheetParseError(cell_address, 'Data type must be time.')
        
    @classmethod
    def parse_worksheet(cls, worksheet):
        validate_cell = \
            lambda cell_address, expected_value: \
            MSR175ShockEvent.validate_cell(worksheet, cell_address, expected_value)

        # Read headers
        validate_cell('A1', 'Event ID:')
        validate_cell('D1', 'Start Date:')
        validate_cell('D2', 'Start Time:')

        event_id   = worksheet['B1'].value
        timestamp  = datetime.combine(MSR175ShockEvent.parse_date(worksheet, 'E1'),
                                      MSR175ShockEvent.parse_time(worksheet, 'E2'))
        
        validate_cell('A6', 'Time (msec)')
        validate_cell('B6', 'X (g)')
        validate_cell('C6', 'Y (g)')
        validate_cell('D6', 'Z (g)')

        sampling_period_ms = None
        prev_t_ms = None
        x_g = []
        y_g = []
        z_g = []
        
        for row in worksheet.iter_rows(min_row = 7, min_col = 1, max_col = 4):
            t_ms = row[0].value
            x_g.append(row[1].value)
            y_g.append(row[2].value)
            z_g.append(row[3].value)
            
            if sampling_period_ms is None and prev_t_ms is None:
                if t_ms != 0.0:
                    message = f'The first time must be 0, but was {t_ms:.5f} ms.'
                    raise MSR175WorksheetParseError(row[0].coordinate, message)
            elif sampling_period_ms is None:
                sampling_period_ms = round(t_ms - prev_t_ms, 6)
            else:
                dt_ms = round(t_ms - prev_t_ms, 6)
                if dt_ms != sampling_period_ms:
                    message = f'The time difference from the previous time must be {sampling_period_ms:.5f} ms, but was {dt_ms:.5f} ms.'
                    raise MSR175WorksheetParseError(row[0].coordinate, message)

            prev_t_ms = t_ms
        
        return MSR175ShockEvent(event_id  = event_id,
                                timestamp = timestamp,
                                sampling_period_ms = sampling_period_ms,
                                x_g = x_g,
                                y_g = y_g,
                                z_g = z_g)

    @staticmethod
    def load_xlsx(xlsx_path, skip_invalid_sheets = True):
        wb     = load_workbook(filename = xlsx_path)
        shock_events = []
        for sheetname in wb.sheetnames:
            worksheet = wb[sheetname]
            
            try:
                shock_event = MSR175ShockEvent.parse_worksheet(worksheet)
                shock_event.__xlsx_path = xlsx_path
                shock_events.append(shock_event)
            except MSR175WorksheetParseError as e:
                cell_address  = e.cell_address
                error_message = e.message

                if skip_invalid_sheets:
                    print(f'''
Warning: Skipped loading data in sheet "{sheetname}" in {xlsx_path}.
Cell "{cell_address}": {error_message}''')
                else:
                    raise MSR175WorkbookLoadError(xlsx_path, sheetname,cell_address, error_message)
                
        return shock_events

def parse_arguments():
    parser = argparse.ArgumentParser(
        description = 'Tool to plot MSR 175 acceleration data.',
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)

    parser.add_argument('--output',
                        dest    = 'output_path',
                        metavar = 'OUTPUT',
                        default = 'output.html',
                        help    = 'Output HTML file path.')
    parser.add_argument('--template',
                        dest    = 'template',
                        metavar = 'TEMPLATE',
                        default = os.path.join(os.path.dirname(__file__), 'template.html'),
                        help    = 'Template for HTML output.')
    parser.add_argument('--skip-invalid-sheets',
                        dest    = 'skip_invalid_sheets',
                        action  = 'store_true',
                        help    = 'Specify this option to skip loading Excel sheets that include invalid format/value and continue processing other Excel sheets.')
    parser.add_argument('--plot-width',
                        dest    = 'plot_width',
                        metavar = 'WIDTH',
                        default = 700,
                        type    = int,
                        help    = 'Width of each plot in pixels.')
    parser.add_argument('--plot-height',
                        dest    = 'plot_height',
                        metavar = 'HEIGHT',
                        default = 350,
                        type    = int,
                        help    = 'Height of each plot in pixels.')
    parser.add_argument('--min-acc',
                        dest    = 'acc_min_g',
                        metavar = 'MIN_G',
                        type    = float,
                        default = float('nan'),
                        help    = 'Minimum acceleration in g for the time series plot. Specify "nan" for auto scale.')
    parser.add_argument('--max-acc',
                        dest    = 'acc_max_g',
                        metavar = 'MAX_G',
                        type    = float,
                        default = float('nan'),
                        help    = 'Maximum acceleration in g for the time series plot. Specify "nan" for auto scale.')
    parser.add_argument('--min-time',
                        dest     = 't_min_ms',
                        type     = float,
                        default  = 0.0,
                        help     = 'Minimum time in the time series plot in milliseconds.')
    parser.add_argument('--max-time',
                        dest     = 't_max_ms',
                        type     = float,
                        default  = float('nan'),
                        help     = 'Maximum time in the time series plot in milliseconds. Specify "nan" for auto scale.')
    parser.add_argument('--min-ps',
                        dest    = 'ps_min_g2',
                        type    = float,
                        default = float('nan'),
                        help    = 'Minimum power spectrum in g^2 for the plot. Specify "nan" for auto scale.')
    parser.add_argument('--max-ps',
                        dest    = 'ps_max_g2',
                        type    = float,
                        default = float('nan'),
                        help    = 'Maximum power spectrum in g^2 for the plot. Specify "nan" for auto scale.')
    parser.add_argument('xlsx_file', nargs='+')

    return parser.parse_args()

def main():
    args = parse_arguments()

    # Read HTML template
    with open(args.template) as template_html_file:
        html_tree = BeautifulSoup(template_html_file, 'html.parser')

    # Check if all Excel filenames are unique.
    # Note: Excel filename (not path) will be used as a part of page anchors, so it must be unique.
    xlsx_paths = [Path(xf) for xf in args.xlsx_file]
    xlsx_filenames = [xlsx_path.name for xlsx_path in xlsx_paths]
    for xlsx_filename in xlsx_filenames:
        if xlsx_filenames.count(xlsx_filename) > 1:
            print(f'ERROR: duplicate Excel file names: {xlsx_filename}', file = sys.stderr)
            exit(1)
    
    # Read Excel files generated by MSR Report Generator to read MSR175 accelerometer data.
    shock_events = []
    for xlsx_path in xlsx_paths:
        shock_events.extend(MSR175ShockEvent.load_xlsx(xlsx_path,
                                                       skip_invalid_sheets = args.skip_invalid_sheets))

    # Generate time series plots for each shock event
    time_series_plots = []
    for shock_event in shock_events:
        plot = shock_event.time_series_plot(width  = args.plot_width,
                                            height = args.plot_height,
                                            acc_min_g = None if np.isnan(args.acc_min_g) else args.acc_min_g,
                                            acc_max_g = None if np.isnan(args.acc_max_g) else args.acc_max_g,
                                            t_min_ms  = None if np.isnan(args.t_min_ms) else args.t_min_ms,
                                            t_max_ms  = None if np.isnan(args.t_max_ms) else args.t_max_ms)
        time_series_plots.append(plot)

    # Generate power spectrum plots for each shock event
    power_spectrum_plots = []
    for shock_event in shock_events:
        plot = shock_event.power_spectrum_plot(width  = args.plot_width,
                                               height = args.plot_height,
                                               ps_min_g2 = None if np.isnan(args.ps_min_g2) else args.ps_min_g2,
                                               ps_max_g2 = None if np.isnan(args.ps_max_g2) else args.ps_max_g2)
        power_spectrum_plots.append(plot)

    # Generate Bokeh JavaScript and "div" tags for plots.
    plots = time_series_plots + power_spectrum_plots
    bokeh_script_html, plot_div_htmls = components(plots)
    plot_divs = [BeautifulSoup(html, 'html.parser') for html in plot_div_htmls]
    time_series_plot_divs = plot_divs[:len(time_series_plots)]
    power_spectrum_plot_divs = plot_divs[len(time_series_plots):]
        
    # Insert Bokeh JavaScript in the HTML tree.
    bokeh_cdn = html_tree.new_tag('script',
                                  src = f'https://cdn.bokeh.org/bokeh/release/bokeh-{bokeh.__version__}.min.js')
    html_tree.head.append(bokeh_cdn)
    
    bokeh_script = BeautifulSoup(bokeh_script_html, 'html.parser')
    html_tree.head.append(bokeh_script)

    # Insert plots.
    plots_container = html_tree.select('#plot-msr175-plots')[0]
    for i in range(len(shock_events)):
        shock_event = shock_events[i]
        time_series_plot_div = time_series_plot_divs[i]
        power_spectrum_plot_div = power_spectrum_plot_divs[i]

        # Add summary.
        summary_table = html_tree.select('table#plot-msr175-summary')[0]
        if i == 0:
            tr = html_tree.new_tag('tr')
            for header_names in ('Data Source',
                                 'Event ID',
                                 'Timestamp',
                                 'Max Acceleration',
                                 'Link'):
                th = html_tree.new_tag('th')
                th.string = header_names
                tr.append(th)
            summary_table.append(tr)

        tr = html_tree.new_tag('tr')
        for cell_content in (shock_event.xlsx_filename,
                             str(shock_event.event_id),
                             shock_event.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                             f'{shock_event.max_g:.1f} g',
                             hyperlink(html_tree, 'jump', f'#{shock_event.html_id}')):
            td = html_tree.new_tag('td')
            if isinstance(cell_content, Tag):
                td.append(cell_content)
            else:
                td.string = cell_content
            tr.append(td)
        summary_table.append(tr)

        # Add title.
        plot_title = html_tree.new_tag('h2', id = shock_event.html_id)
        plot_title.string = f'{shock_event.xlsx_filename}: Event ID {shock_event.event_id}'
        plots_container.append(plot_title)

        # Add event properties.
        prop_table = property_table(html_tree,
                                    ('Data Source:',
                                     'Event ID:',
                                     'Timestamp:',
                                     'Max Acceleration:'),
                                    (shock_event.xlsx_filename,
                                     str(shock_event.event_id),
                                     shock_event.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                                     f'{shock_event.max_g:.1f} g'))
        plots_container.append(prop_table)

        # Add plots
        horizontal_container = html_tree.new_tag('div')
        horizontal_container.append(time_series_plot_div)
        horizontal_container.append(power_spectrum_plot_div)
        plots_container.append(horizontal_container)

    # Output the final HTML.
    with open(args.output_path, 'w') as output_html_file:
        output_html_file.write(html_tree.prettify())
        print(f'Generated {args.output_path}.')

if __name__ == "__main__":
    main()
